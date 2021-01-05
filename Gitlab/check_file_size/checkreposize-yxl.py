#!/usr/bin/python
#coding=utf-8
import xlwt
import sys
import ssl
import string
from urllib.parse import quote
import urllib.request
from urllib import error
import os
import re
import openpyxl
ssl._create_default_https_context = ssl._create_unverified_context
TOKEN = "&private_token"
IP = "https://git.com/api/v3"
global false, null, true

false = null = true = ''

def Gitlab_api(IP,TOKEN,API):
	url = IP + API + TOKEN
	#url = quote(url, safe=string.printable)
        #paichu = string.printable.replace(" ","")
        #paichu = paichu.replace("+","")
	#print (url)
	url = quote(url,safe='/:?=&')
	try:
		res = urllib.request.urlopen(url)
	except error.URLError as e:
		print (str(url)+"is error!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
		print (e.reason)
	else:
		response = res.read()
		response = eval(response)
		return response


def hum_convert(value):
    units = ["B", "KB", "MB", "GB", "TB", "PB"]
    size = 1024.0
    for i in range(len(units)):
        if (value / size) < 1:
              return "%.2f%s" % (value, units[i])
        value = value / size

######




outwb = openpyxl.Workbook()
sheet1 = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
sheet2 = outwb.create_sheet(index=1)  # 在将写的文件创建sheet

data1 = ["仓库","文件名","大小","字节大小"]
data2 = ["仓库","大小","字节大小"]
for x in range(0,len(data1)):
	sheet1.cell(column = x+1 , row = 1 , value = data1[x]) #表头
for x in range(0,len(data2)):
	sheet2.cell(column = x+1 , row = 1 , value = data2[x]) #表头




hangshu=2
all_hangshu=2
number=0
project_id_dict = {} #all project name and id eg: {"cm/jenkins-pipeline":"2009"}
#project = Gitlab_api(IP,TOKEN,"/projects?simple=true")
#project = Gitlab_api(IP,TOKEN,"/projects?all=true")
commond = "curl -s --head -k https://git.com/api/v3/projects/all?per_page=100\&order_by=id\&private_token=rsuRSng | grep X-Total-Pages"
result = os.popen(commond).readlines()
result = re.findall(r"\d+",str(result))
#print (result)
result = "".join(result)
for P in range(0,int(result)):
	P = P + 1
	project = Gitlab_api(IP,TOKEN,"/projects/all?per_page=100&page=%s&order_by=id" % P)
	for Q in project:
		#print (Q)
		project_id_dict[Q['path_with_namespace']] = Q['id']

######
#print (len(project_id_dict.keys()))

for i in project_id_dict.keys():
	print ("第几个项目:",all_hangshu)
	#print (i)
	repository_file = Gitlab_api(IP,TOKEN,"/projects/%s/repository/tree?recursive=true&ref_name=master"%(project_id_dict[i]))
	#print repository_file
	project_size = 0
	#print (type(repository_file))
	if repository_file is None:
		#print ("跳出循环")
		continue
	for file in repository_file:
		if  file['mode'] == "100644":
			file_size = Gitlab_api(IP,TOKEN,"/projects/%s/repository/files?file_path=%s&ref=master"%(project_id_dict[i],file['path'])) 
			if file_size is None:
				continue
			file_size_read = hum_convert(file_size['size'])	
			data_list = []
			data_list.append(i)
			data_list.append(file['path'])
			data_list.append(file_size_read)
			data_list.append(file_size['size'])
			
			
			if file_size['size'] > 52428800:
				print ("开始写入",i,file['path'],file_size_read,file_size['size'])
				for c in range(0,len(data1)):
					sheet1.cell(column = c+1 , row = hangshu , value = data_list[c])
				hangshu = hangshu + 1


		
			project_size = project_size + file_size['size']
	
	all_data_list = []
	all_data_list.append(i)
	all_data_list.append(hum_convert(project_size))
	all_data_list.append(project_size)
	

	for H in range(0,len(data2)):
		sheet2.cell(column = H+1,row = all_hangshu,value = all_data_list[H]) 


	all_hangshu = all_hangshu + 1
	#if number == 20:
	#savexlsx = "./results.xlsx"
	#	outwb.save(savexlsx)  # 保存结果
	#	print ("done")
	#	sys.exit(1)
	#else:
	#	number = number + 1
savexlsx = "./results.xlsx"
outwb.save(savexlsx)  # 保存结果
print ("done")
